import xlrd
from xlrd import xldate_as_tuple
import openpyxl
import datetime


class ExcelData():
    def __init__(self, file_path, sheet_name):
        self.file_path = file_path
        self.sheet_name = sheet_name
        self.workbook = xlrd.open_workbook(self.file_path)

        # 获取工作表的内容
        self.table = self.workbook.sheet_by_name(self.sheet_name)
        # 获取第一行内容
        self.keys = self.table.row_values(0)
        # 获取行数
        self.rowNum = self.table.nrows
        # 获取列数
        self.colNum = self.table.ncols

    def readExcel(self):
        datas = []
        for i in range(1, self.rowNum):
            sheet_data = []
            for j in range(self.colNum):
                # 获取单元格类型
                c_type = self.table.cell(i, j).ctype
                # 获取单元格数据
                c_cell = self.table.cell_value(i, j)
                if c_type == 2 and c_cell % 1 == 0:
                    c_cell = int(c_cell)
                elif c_type == 3:
                    date = datetime.datetime(*xldate_as_tuple(c_cell, 0))
                    c_cell = date.strftime('%Y/%d/%m %H:%M:%S')
                elif c_type == 4:
                    c_cell = True if c_cell == 1 else False
                # sheet_data[self.keys[j]] = c_cell   # 字典
                sheet_data.append(c_cell)
            datas.append(sheet_data)
        return datas

    def write(self, rowNum, colNum, result):
        workbook = openpyxl.load_workbook(self.file_path)
        table = workbook.get_sheet_by_name(self.sheet_name)
        table = workbook.active

        # rows = table.max_row
        # cols = table.max_column
        # values = ['E','X','C','E','L']
        # for value in values:
        #     table.cell(rows + 1, 1).value = value
        #     rows = rows + 1

        # 指定单元格中写入数据
        table.cell(rowNum, colNum, result)
        workbook.save(self.file_path)


if __name__ == '__main__':
    file_path = "D:\python_data\接口自动化测试.xlsx"
    sheet_name = "测试用例"
    data = ExcelData(file_path, sheet_name)
    datas = data.readExcel()
    print(datas)
    print(type(datas))
    for i in datas:
        print(i)

    # data.write(2,12,"哈哈")
